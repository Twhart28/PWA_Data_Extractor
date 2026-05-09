from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment

# Analysis mode selector. Choose 1 to use combined peripheral SYS/DIA/MEAN
# matching or 2 to match only on peripheral systolic pressure.
ANALYSIS_MODE = 2

APP_DIR = Path(__file__).resolve().parent
APP_TITLE = "PWA Data Extractor"
APP_VERSION = "2.0.0"
APP_PUBLISHER = "Thomas Hart"
APP_SUBTITLE = (
    "Process PWA reports, review multi-entry patients, and export the "
    "same Excel workbook structure."
)
APP_ICON_PATH = APP_DIR / "App_Logo.ico"
README_PATH = APP_DIR / "README.md"
CONTACT_EMAIL = "thomaswhart28@gmail.com"
REPOSITORY_URL = "https://github.com/Twhart28/PWA_Data_Extractor"

# Report parsing modes (user-selected per batch).
REPORT_MODE_DETAILED = "detailed"
REPORT_MODE_CLINICAL = "clinical"

# Grouping modes — drive how filename segments build the Patient ID.
GROUP_MODE_SUBJECT = "subject"
GROUP_MODE_SUBJECT_TIMEPOINT = "subject_timepoint"
GROUP_MODE_SUBJECT_VISIT = "subject_visit"
GROUP_MODE_SUBJECT_VISIT_TIMEPOINT = "subject_visit_timepoint"

COLUMNS = [
    "Source File",
    "Patient ID",
    "Subject ID",
    "Visit",
    "Timepoint",
    "Report Type",
    "Scanned ID",
    "Scan Date",
    "Scan Time",
    "Record #",
    "Analyed",
    "Date of Birth",
    "Age",
    "Gender",
    "Height (m)",
    "# of Pulses",
    "Pulse Height",
    "Pulse Height Variation (%)",
    "Diastolic Variation (%)",
    "Shape Deviation (%)",
    "Pulse Length Variation (%)",
    "Overall Quality (%)",
    "Peripheral Systolic Pressure (mmHg)",
    "Peripheral Diastolic Pressure (mmHg)",
    "Peripheral Pulse Pressure (mmHg)",
    "Peripheral Mean Pressure (mmHg)",
    "Aortic Systolic Pressure (mmHg)",
    "Aortic Diastolic Pressure (mmHg)",
    "Aortic Pulse Pressure (mmHg)",
    "Heart Rate (bpm)",
    "Pulse Pressure Amplification (%)",
    "Period (ms)",
    "Ejection Duration (ms)",
    "Ejection Duration (%)",
    "Aortic T2 (ms)",
    "P1 Height (mmHg)",
    "Aortic Augmentation (mmHg)",
    "Aortic AIx AP/PP(%)",
    "Aortic AIx P2/P1(%)",
    "Aortic AIx AP/PP @ HR75 (%)",
    "Buckberg SEVR (%)",
    "PTI Systolic (mmHg.s/min)",
    "PTI Diastolic (mmHg.s/min)",
    "End Systolic Pressure (mmHg)",
    "MAP Systolic (mmHg)",
    "MAP Diastolic (mmHg)",
]

EXTRA_COLUMNS = ["Source Path"]
ALL_DATA_COLUMNS = [*COLUMNS, *EXTRA_COLUMNS]
UI_CONTEXT_COLUMNS = ["Subject ID", "Visit", "Timepoint", "Report Type"]
WORKBOOK_COLUMNS = [column for column in COLUMNS if column not in UI_CONTEXT_COLUMNS]
CLINICAL_WORKBOOK_COLUMNS = [
    "Source File",
    "Patient ID",
    "Scanned ID",
    "Scan Date",
    "Scan Time",
    "Record #",
    "Analyed",
    "Date of Birth",
    "Age",
    "Gender",
    "Height (m)",
    "# of Pulses",
    "Peripheral Systolic Pressure (mmHg)",
    "Peripheral Diastolic Pressure (mmHg)",
    "Peripheral Pulse Pressure (mmHg)",
    "Peripheral Mean Pressure (mmHg)",
    "Aortic Systolic Pressure (mmHg)",
    "Aortic Diastolic Pressure (mmHg)",
    "Aortic Pulse Pressure (mmHg)",
    "Heart Rate (bpm)",
]
DETAILED_REPORT_MARKER = "PWA Detailed Report"
CLINICAL_REPORT_MARKER = "PWA Clinical Report"
CLINICAL_REPORT_MESSAGE = (
    "Recognized as a Clinical Report — switch report type to Clinical to include it"
)
DETAILED_REPORT_MESSAGE = (
    "Recognized as a Detailed Report — switch report type to Detailed to include it"
)
UNRECOGNIZED_REPORT_MESSAGE = "Not recognized as a PWA report"
SPECIAL_ROW_MESSAGES = frozenset(
    {
        CLINICAL_REPORT_MESSAGE,
        DETAILED_REPORT_MESSAGE,
        UNRECOGNIZED_REPORT_MESSAGE,
    }
)
ANALYSIS_FIELDS_BY_MODE: dict[int, list[str]] = {
    1: [
        "Peripheral Systolic Pressure (mmHg)",
        "Peripheral Diastolic Pressure (mmHg)",
        "Peripheral Mean Pressure (mmHg)",
    ],
    2: ["Peripheral Systolic Pressure (mmHg)"],
}
PAIR_DIFF_SOURCE_FIELDS = [
    "Peripheral Systolic Pressure (mmHg)",
    "Peripheral Diastolic Pressure (mmHg)",
    "Peripheral Mean Pressure (mmHg)",
    "Aortic Systolic Pressure (mmHg)",
    "Aortic Diastolic Pressure (mmHg)",
]
PAIR_DIFF_EXPORT_COLUMNS = {
    "Peripheral Systolic Pressure (mmHg)": "Pair Diff Peripheral Systolic (mmHg)",
    "Peripheral Diastolic Pressure (mmHg)": "Pair Diff Peripheral Diastolic (mmHg)",
    "Peripheral Mean Pressure (mmHg)": "Pair Diff Peripheral Mean (mmHg)",
    "Aortic Systolic Pressure (mmHg)": "Pair Diff Aortic Systolic (mmHg)",
    "Aortic Diastolic Pressure (mmHg)": "Pair Diff Aortic Diastolic (mmHg)",
}

# Review reasons surfaced to the UI.
REVIEW_REASON_MULTI_ENTRY = "multi_entry"
REVIEW_REASON_PAIR_ALERT = "pair_alert"
REVIEW_REASON_BOTH = "multi_entry_pair_alert"

README_FALLBACK_TEXT = """# PWA Data Extractor

Extract PWA Detailed or Clinical report PDFs, review repeated measurements, and export a structured Excel workbook.

## Workflow

1. Add PWA PDF reports.
2. Choose Detailed or Clinical mode.
3. Choose filename grouping and optionally paste an AI-generated regex.
4. Process the reports locally.
5. Review flagged pairs and confirm the selected rows.
6. Export the workbook.

## Output

The export contains:

- **All Data**
- **Kept Data**
- **Averaged Data**
- **Skipped Files**

Skipped Files includes one-file subjects, wrong report types, and unrecognized PDFs.
"""


@dataclass
class FilenameMetadata:
    subject_id: str | None
    visit: str | None
    timepoint: str | None
    grouping_key: str


@dataclass
class ReviewItem:
    patient_id: str
    reason: str  # one of REVIEW_REASON_*


@dataclass
class AnalysisBundle:
    dataframe: pd.DataFrame
    special_row_mask: pd.Series
    analyzed_df: pd.DataFrame
    kept_indices: set[int]
    used_pairs: dict[str, tuple[int, int]]
    auto_pairs: dict[str, tuple[int, int]]
    manual_patients: list[str]
    review_items: list[ReviewItem] = field(default_factory=list)


def load_readme_text() -> str:
    if README_PATH.exists():
        return README_PATH.read_text(encoding="utf-8")
    return README_FALLBACK_TEXT


def default_output_path() -> Path:
    timestamp = datetime.now().strftime("%m-%d-%y %H-%M")
    return Path.home() / "Downloads" / f"PWA Export ({timestamp}).xlsx"


def extract_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = [page.extract_text() or "" for page in pdf.pages]
    return "\n".join(pages_text)


def _search(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return match.group(1) if match else None


def _to_number(value: str) -> int | float | str:
    normalized = value.strip()
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", normalized):
        return float(normalized) if "." in normalized else int(normalized)
    return value


def _extract_scan_datetime(text: str) -> tuple[str | None, str | None]:
    labeled_match = re.search(
        r"Date(?:\s+and)?\s+Time:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})\s+([0-9]{2}:[0-9]{2}(?::[0-9]{2})?)",
        text,
        flags=re.IGNORECASE,
    )
    if labeled_match:
        return labeled_match.group(1), labeled_match.group(2)

    date_time_match = None
    for date_time_match in re.finditer(
        r"([0-9]{2}/[0-9]{2}/[0-9]{4})\s+([0-9]{2}:[0-9]{2}(?::[0-9]{2})?)",
        text,
    ):
        pass
    if date_time_match:
        return date_time_match.group(1), date_time_match.group(2)
    return None, None


def _clean_component(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"[^A-Za-z0-9]+", "", value).strip()
    return cleaned.upper() if cleaned else None


def _extract_numeric_component(value: str | None) -> str | None:
    cleaned = _clean_component(value)
    if not cleaned:
        return None
    match = re.search(r"(\d+)", cleaned)
    return match.group(1) if match else None


def parse_filename_metadata(
    pdf_path: Path,
    grouping_mode: str = GROUP_MODE_SUBJECT,
    filename_pattern: str | None = None,
) -> FilenameMetadata:
    """Parse Subject / Visit / Timepoint segments out of a PDF filename.

    The grouping_key is built per the selected mode using a space-separated
    format with V<n> and T<n> prefixes (e.g. ``01 T000`` or ``01 V2 T0``).
    """
    stem = pdf_path.stem.strip()
    if not stem:
        return FilenameMetadata(None, None, None, "")

    def _build_grouping_key(
        resolved_subject: str | None,
        resolved_visit: str | None,
        resolved_timepoint: str | None,
    ) -> str:
        key_parts: list[str] = []
        if resolved_subject:
            key_parts.append(resolved_subject)

        if grouping_mode == GROUP_MODE_SUBJECT_VISIT and resolved_visit:
            key_parts.append(f"V{resolved_visit}")
        elif grouping_mode == GROUP_MODE_SUBJECT_TIMEPOINT and resolved_timepoint:
            key_parts.append(f"T{resolved_timepoint}")
        elif grouping_mode == GROUP_MODE_SUBJECT_VISIT_TIMEPOINT:
            if resolved_visit:
                key_parts.append(f"V{resolved_visit}")
            if resolved_timepoint:
                key_parts.append(f"T{resolved_timepoint}")

        return " ".join(key_parts) if key_parts else stem.strip()

    if filename_pattern:
        try:
            custom_match = re.search(filename_pattern, stem, flags=re.IGNORECASE)
        except re.error:
            custom_match = None
        if custom_match and "subject" in custom_match.groupdict():
            custom_subject = _clean_component(custom_match.group("subject"))
            if custom_subject:
                custom_visit = _clean_component(custom_match.groupdict().get("visit"))
                custom_timepoint = _clean_component(
                    custom_match.groupdict().get("timepoint")
                )
                return FilenameMetadata(
                    custom_subject,
                    custom_visit,
                    custom_timepoint,
                    _build_grouping_key(
                        custom_subject,
                        custom_visit,
                        custom_timepoint,
                    ),
                )

    tokens = [token for token in re.split(r"[\s_-]+", stem) if token]
    subject_id: str | None = None
    visit: str | None = None
    timepoint: str | None = None
    subject_index = -1

    visit_token_re = re.compile(r"V(?:ISIT)?\d+", re.IGNORECASE)
    timepoint_token_re = re.compile(r"(?:T|TP|TIMEPOINT)\d+", re.IGNORECASE)
    ignored_suffix_re = re.compile(r"(?:PWA|REPORT|RPT|RUN|MEAS|MEASURE)\d*$", re.IGNORECASE)
    compact_timepoint_re = re.compile(r"\d+[A-Za-z]?$")

    def _is_visit_or_timepoint_token(value: str) -> bool:
        return bool(visit_token_re.fullmatch(value) or timepoint_token_re.fullmatch(value))

    def _is_implicit_timepoint_candidate(value: str) -> bool:
        cleaned = _clean_component(value)
        if not cleaned or ignored_suffix_re.fullmatch(cleaned):
            return False
        return bool(compact_timepoint_re.fullmatch(cleaned))

    # Prefer tokens that look like "IAS003", "S01", "subj01", etc.
    # Preserve the prefix because it can carry the study identifier.
    # Skip explicit visit/timepoint tokens like "V2" or "T0".
    for index, token in enumerate(tokens):
        if _is_visit_or_timepoint_token(token):
            continue
        if re.fullmatch(r"[A-Za-z]+[A-Za-z0-9]*\d+[A-Za-z0-9]*", token):
            subject_id = _clean_component(token)
            subject_index = index
            break

    # Fall back to the first token containing any digits (still skipping V/T).
    if subject_id is None:
        for index, token in enumerate(tokens):
            if _is_visit_or_timepoint_token(token):
                continue
            if re.search(r"\d", token):
                subject_id = _extract_numeric_component(token) or _clean_component(token)
                subject_index = index
                break

    search_tokens = tokens[subject_index + 1 :] if subject_index >= 0 else tokens
    numeric_candidates: list[str] = []
    explicit_visit: str | None = None
    explicit_timepoint: str | None = None

    for token in search_tokens:
        cleaned_token = token.strip()

        visit_match = re.fullmatch(r"V(?:ISIT)?(\d+)", cleaned_token, flags=re.IGNORECASE)
        if visit_match:
            explicit_visit = visit_match.group(1)
            continue

        timepoint_match = re.fullmatch(
            r"(?:T|TP|TIMEPOINT)(\d+)",
            cleaned_token,
            flags=re.IGNORECASE,
        )
        if timepoint_match:
            explicit_timepoint = timepoint_match.group(1)
            continue

        numeric_value = _extract_numeric_component(cleaned_token)
        if numeric_value is not None and _is_implicit_timepoint_candidate(cleaned_token):
            numeric_candidates.append(numeric_value)

    visit = explicit_visit
    timepoint = explicit_timepoint

    # Heuristic: if neither V/T prefix was found, treat trailing numeric tokens
    # as visit/timepoint based on order.
    if visit is None and timepoint is None:
        if len(numeric_candidates) >= 2:
            visit = numeric_candidates[0]
            timepoint = numeric_candidates[1]
        elif len(numeric_candidates) == 1:
            timepoint = numeric_candidates[0]
    else:
        remaining_candidates = list(numeric_candidates)
        if visit is None and remaining_candidates:
            visit = remaining_candidates.pop(0)
        if timepoint is None and remaining_candidates:
            timepoint = remaining_candidates.pop(0)

    fallback_subject = _extract_numeric_component(stem) or _clean_component(stem) or stem.strip()
    resolved_subject = subject_id or fallback_subject

    grouping_key = _build_grouping_key(resolved_subject, visit, timepoint)
    return FilenameMetadata(subject_id, visit, timepoint, grouping_key)


def _finalize_record(record: dict[str, object]) -> dict[str, object]:
    for key, value in record.items():
        if isinstance(value, str):
            record[key] = _to_number(value)
    return record


def parse_detailed_report_text(text: str) -> dict[str, object]:
    normalized = re.sub(r"\s+", " ", text)

    patient_id = _search(r"Patient ID:\s*(\S+)", normalized)
    dob = _search(r"Date Of Birth:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", normalized)
    scan_date, scan_time = _extract_scan_datetime(normalized)

    age_gender_match = re.search(
        r"Age, Gender:\s*([0-9]+),\s*([A-Za-z]+)",
        normalized,
        flags=re.IGNORECASE,
    )
    age = age_gender_match.group(1) if age_gender_match else None
    gender = age_gender_match.group(2) if age_gender_match else None

    height_cm = _search(r"Height:\s*([0-9.]+)\s*cm", normalized)
    height_m = round(float(height_cm) / 100, 2) if height_cm else None

    pulses = _search(r"Number Of Pulses:\s*([0-9]+)", normalized)

    heart_rate_period = re.search(
        r"Heart Rate, Period:\s*([0-9.]+)\s*bpm,\s*([0-9.]+)\s*ms",
        normalized,
        flags=re.IGNORECASE,
    )
    heart_rate = heart_rate_period.group(1) if heart_rate_period else None
    period = heart_rate_period.group(2) if heart_rate_period else None

    ejection_match = re.search(
        r"Ejection Duration \(ED\):\s*([0-9.]+)\s*ms,\s*([0-9.]+)\s*%",
        normalized,
        flags=re.IGNORECASE,
    )
    ejection_ms = ejection_match.group(1) if ejection_match else None
    ejection_pct = ejection_match.group(2) if ejection_match else None

    aortic_t2 = _search(r"Aortic T2:\s*([0-9.]+)\s*ms", normalized)
    p1_height = _search(r"P1 Height.*?:\s*([0-9.]+)\s*mmHg", normalized)
    aortic_augmentation = _search(
        r"Aortic Augmentation.*?:\s*([-+]?[0-9.]+)\s*mmHg",
        normalized,
    )

    aix_match = re.search(
        r"Aortic AIx \(AP/PP, P2/P1\):\s*([-+]?[0-9.]+)\s*%,\s*([-+]?[0-9.]+)\s*%",
        normalized,
        flags=re.IGNORECASE,
    )
    aortic_aix_ap_pp = aix_match.group(1) if aix_match else None
    aortic_aix_p2_p1 = aix_match.group(2) if aix_match else None

    aix_hr75 = _search(
        r"Aortic AIx \(AP/PP\) @HR75:\s*([-+]?[0-9.]+)\s*%",
        normalized,
    )
    buckberg = _search(r"Buckberg SEVR:\s*([0-9.]+)\s*%", normalized)

    pti_match = re.search(
        r"PTI \(Systole, Diastole\):\s*([0-9.]+),\s*([0-9.]+)\s*mmHg\.s/min",
        normalized,
        flags=re.IGNORECASE,
    )
    pti_systolic = pti_match.group(1) if pti_match else None
    pti_diastolic = pti_match.group(2) if pti_match else None

    end_systolic_pressure = _search(
        r"End Systolic Pressure:\s*([0-9.]+)\s*mmHg",
        normalized,
    )

    map_match = re.search(
        r"MAP \(Systole, Diastole\):\s*([0-9.]+),\s*([0-9.]+)\s*mmHg",
        normalized,
        flags=re.IGNORECASE,
    )
    map_systolic = map_match.group(1) if map_match else None
    map_diastolic = map_match.group(2) if map_match else None

    pulse_height = _search(r"Pulse Height:\s*([0-9.]+)", normalized)
    pulse_height_variation = _search(
        r"Pulse Height Variation:\s*([0-9.]+)\s*%",
        normalized,
    )
    diastolic_variation = _search(
        r"Diastolic Variation:\s*([0-9.]+)\s*%",
        normalized,
    )
    shape_deviation = _search(r"Shape Deviation:\s*([0-9.]+)\s*%", normalized)
    pulse_length_variation = _search(
        r"Pulse Length Variation:\s*([0-9.]+)\s*%",
        normalized,
    )
    overall_quality = _search(r"Overall Quality:\s*([0-9.]+)\s*%", normalized)

    amplification = _search(r"PP Amplification:\s*([0-9.]+)\s*%", normalized)

    brachial_match = re.search(
        r"Brachial SYS/DIA:\s*([0-9.]+)/([0-9.]+)",
        normalized,
        flags=re.IGNORECASE,
    )
    peripheral_sys = brachial_match.group(1) if brachial_match else None
    peripheral_dia = brachial_match.group(2) if brachial_match else None

    aortic_sys = None
    aortic_dia = None
    peripheral_pp = None
    aortic_pp = None
    peripheral_mean = None
    table_heart_rate = None

    sp_match = re.search(r"SP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if sp_match:
        peripheral_sys = peripheral_sys or sp_match.group(1)
        aortic_sys = sp_match.group(2)

    dp_match = re.search(r"DP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if dp_match:
        peripheral_dia = peripheral_dia or dp_match.group(1)
        aortic_dia = dp_match.group(2)

    pp_match = re.search(r"PP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if pp_match:
        peripheral_pp = pp_match.group(1)
        aortic_pp = pp_match.group(2)

    map_hr_match = re.search(
        r"MAP HR\s+([0-9.]+)\s+([0-9.]+)",
        normalized,
        flags=re.IGNORECASE,
    )
    if map_hr_match:
        peripheral_mean = map_hr_match.group(1)
        table_heart_rate = map_hr_match.group(2)

    if peripheral_sys and peripheral_dia and peripheral_pp is None:
        try:
            peripheral_pp = str(float(peripheral_sys) - float(peripheral_dia))
        except ValueError:
            peripheral_pp = None

    if aortic_sys and aortic_dia and aortic_pp is None:
        try:
            aortic_pp = str(float(aortic_sys) - float(aortic_dia))
        except ValueError:
            aortic_pp = None

    heart_rate = heart_rate or table_heart_rate

    record = {
        "Scanned ID": patient_id,
        "Scan Date": scan_date,
        "Scan Time": scan_time,
        "Date of Birth": dob,
        "Age": age,
        "Gender": gender,
        "Height (m)": height_m,
        "# of Pulses": pulses,
        "Pulse Height": pulse_height,
        "Pulse Height Variation (%)": pulse_height_variation,
        "Diastolic Variation (%)": diastolic_variation,
        "Shape Deviation (%)": shape_deviation,
        "Pulse Length Variation (%)": pulse_length_variation,
        "Overall Quality (%)": overall_quality,
        "Peripheral Systolic Pressure (mmHg)": peripheral_sys,
        "Peripheral Diastolic Pressure (mmHg)": peripheral_dia,
        "Peripheral Pulse Pressure (mmHg)": peripheral_pp,
        "Peripheral Mean Pressure (mmHg)": peripheral_mean,
        "Aortic Systolic Pressure (mmHg)": aortic_sys,
        "Aortic Diastolic Pressure (mmHg)": aortic_dia,
        "Aortic Pulse Pressure (mmHg)": aortic_pp,
        "Heart Rate (bpm)": heart_rate,
        "Pulse Pressure Amplification (%)": amplification,
        "Period (ms)": period,
        "Ejection Duration (ms)": ejection_ms,
        "Ejection Duration (%)": ejection_pct,
        "Aortic T2 (ms)": aortic_t2,
        "P1 Height (mmHg)": p1_height,
        "Aortic Augmentation (mmHg)": aortic_augmentation,
        "Aortic AIx AP/PP(%)": aortic_aix_ap_pp,
        "Aortic AIx P2/P1(%)": aortic_aix_p2_p1,
        "Aortic AIx AP/PP @ HR75 (%)": aix_hr75,
        "Buckberg SEVR (%)": buckberg,
        "PTI Systolic (mmHg.s/min)": pti_systolic,
        "PTI Diastolic (mmHg.s/min)": pti_diastolic,
        "End Systolic Pressure (mmHg)": end_systolic_pressure,
        "MAP Systolic (mmHg)": map_systolic,
        "MAP Diastolic (mmHg)": map_diastolic,
    }

    return _finalize_record(record)


def parse_clinical_report_text(text: str) -> dict[str, object]:
    """Parse a PWA Clinical Report's smaller field set."""
    normalized = re.sub(r"\s+", " ", text)

    scanned_id = _search(r"Patient ID:\s*(\S+)", normalized)
    dob = _search(r"Date Of Birth:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", normalized)
    scan_date, scan_time = _extract_scan_datetime(normalized)

    age_gender_match = re.search(
        r"Age,\s*Gender:\s*([0-9]+),\s*([A-Za-z]+)",
        normalized,
        flags=re.IGNORECASE,
    )
    age = age_gender_match.group(1) if age_gender_match else None
    gender = age_gender_match.group(2) if age_gender_match else None

    height_cm = _search(r"Height:\s*([0-9.]+)\s*cm", normalized)
    height_m = round(float(height_cm) / 100, 2) if height_cm else None
    pulses = _search(r"Number Of Pulses:\s*([0-9]+)", normalized)

    brachial_match = re.search(
        r"Brachial SYS/DIA:\s*([0-9.]+)\s*/\s*([0-9.]+)\s*mmHg",
        normalized,
        flags=re.IGNORECASE,
    )
    peripheral_sys = brachial_match.group(1) if brachial_match else None
    peripheral_dia = brachial_match.group(2) if brachial_match else None

    aortic_sys = _search(r"Aortic SP:\s*([0-9.]+)\s*mmHg", normalized)
    aortic_dia = _search(r"\bDP:\s*([0-9.]+)\s*mmHg", normalized)
    aortic_pp = _search(r"\bPP:\s*([0-9.]+)\s*mmHg", normalized)
    heart_rate = _search(r"\bHR:\s*([0-9.]+)\s*bpm", normalized)

    peripheral_pp = None
    peripheral_mean = None
    if peripheral_sys is not None and peripheral_dia is not None:
        try:
            peripheral_sys_value = float(peripheral_sys)
            peripheral_dia_value = float(peripheral_dia)
            peripheral_pp = peripheral_sys_value - peripheral_dia_value
            peripheral_mean = round(
                (peripheral_sys_value + (2 * peripheral_dia_value)) / 3,
                2,
            )
        except ValueError:
            peripheral_pp = None
            peripheral_mean = None

    record = {
        "Scanned ID": scanned_id,
        "Scan Date": scan_date,
        "Scan Time": scan_time,
        "Date of Birth": dob,
        "Age": age,
        "Gender": gender,
        "Height (m)": height_m,
        "# of Pulses": pulses,
        "Peripheral Systolic Pressure (mmHg)": peripheral_sys,
        "Peripheral Diastolic Pressure (mmHg)": peripheral_dia,
        "Peripheral Pulse Pressure (mmHg)": peripheral_pp,
        "Peripheral Mean Pressure (mmHg)": peripheral_mean,
        "Aortic Systolic Pressure (mmHg)": aortic_sys,
        "Aortic Diastolic Pressure (mmHg)": aortic_dia,
        "Aortic Pulse Pressure (mmHg)": aortic_pp,
        "Heart Rate (bpm)": heart_rate,
    }

    return _finalize_record(record)


def detect_report_type(text: str) -> str:
    normalized = text.lower()
    if DETAILED_REPORT_MARKER.lower() in normalized:
        return REPORT_MODE_DETAILED
    if CLINICAL_REPORT_MARKER.lower() in normalized:
        return REPORT_MODE_CLINICAL
    return "unrecognized"


def empty_record(message: str, pdf_path: Path) -> dict[str, object]:
    record: dict[str, object] = {column: None for column in COLUMNS}
    record["Source File"] = pdf_path.name
    record["Source Path"] = str(pdf_path)
    record["Patient ID"] = message
    return record


def _apply_filename_metadata(
    record: dict[str, object],
    pdf_path: Path,
    report_type: str,
    grouping_mode: str,
    filename_pattern: str | None = None,
) -> dict[str, object]:
    metadata = parse_filename_metadata(
        pdf_path,
        grouping_mode=grouping_mode,
        filename_pattern=filename_pattern,
    )
    record["Source File"] = pdf_path.name
    record["Source Path"] = str(pdf_path)
    record["Patient ID"] = metadata.grouping_key or pdf_path.stem.strip() or pdf_path.name
    record["Subject ID"] = metadata.subject_id
    record["Visit"] = metadata.visit
    record["Timepoint"] = metadata.timepoint
    record["Report Type"] = report_type.title()
    return record


def process_pdf(
    pdf_path: Path,
    report_mode: str = REPORT_MODE_DETAILED,
    grouping_mode: str = GROUP_MODE_SUBJECT,
    filename_pattern: str | None = None,
) -> dict[str, object]:
    """Read a PDF and return a record matching the requested report mode.

    PDFs whose content doesn't match ``report_mode`` come back as special rows
    so they remain visible in the workbook without being averaged.
    """
    text = extract_text(pdf_path)
    report_type = detect_report_type(text)

    if report_type == REPORT_MODE_DETAILED:
        if report_mode != REPORT_MODE_DETAILED:
            return empty_record(DETAILED_REPORT_MESSAGE, pdf_path)
        data = parse_detailed_report_text(text)
        return _apply_filename_metadata(
            data,
            pdf_path,
            report_type,
            grouping_mode,
            filename_pattern=filename_pattern,
        )

    if report_type == REPORT_MODE_CLINICAL:
        if report_mode != REPORT_MODE_CLINICAL:
            return empty_record(CLINICAL_REPORT_MESSAGE, pdf_path)
        data = parse_clinical_report_text(text)
        return _apply_filename_metadata(
            data,
            pdf_path,
            report_type,
            grouping_mode,
            filename_pattern=filename_pattern,
        )

    return empty_record(UNRECOGNIZED_REPORT_MESSAGE, pdf_path)


def prepare_dataframe(records: list[dict[str, object]]) -> tuple[pd.DataFrame, pd.Series]:
    df = pd.DataFrame(records)

    for column in ALL_DATA_COLUMNS:
        if column not in df.columns:
            df[column] = None

    df = df[ALL_DATA_COLUMNS]
    df["Special Row"] = df["Patient ID"].isin(SPECIAL_ROW_MESSAGES)
    df.loc[df["Special Row"], COLUMNS[2:]] = None

    df.sort_values(
        by=["Special Row", "Patient ID", "Scan Date", "Scan Time"],
        inplace=True,
    )

    special_rows = df["Special Row"]
    regular_rows = df.loc[~special_rows]
    duplicate_key = ["Patient ID", "Scan Time", "PTI Diastolic (mmHg.s/min)"]
    complete_duplicate_key = regular_rows[duplicate_key].notna().all(axis=1)
    regular_df = pd.concat(
        [
            regular_rows.loc[complete_duplicate_key].drop_duplicates(
                subset=duplicate_key,
                keep="first",
            ),
            regular_rows.loc[~complete_duplicate_key],
        ]
    )
    df = pd.concat([regular_df, df.loc[special_rows]], ignore_index=True)

    df.sort_values(
        by=["Special Row", "Patient ID", "Scan Date", "Scan Time"],
        inplace=True,
        ignore_index=True,
    )

    special_row_mask = df["Special Row"].copy()
    df["Record #"] = None
    valid_rows = ~df["Special Row"]
    df.loc[valid_rows, "Record #"] = (
        df[valid_rows].groupby("Patient ID").cumcount() + 1
    )

    return df, special_row_mask


def closest_pair_indices(
    df: pd.DataFrame,
    fields: list[str],
) -> tuple[int, int] | None:
    if len(df) < 2:
        return None

    systolic_only = fields == ["Peripheral Systolic Pressure (mmHg)"]
    diastolic_values = (
        pd.to_numeric(df["Peripheral Diastolic Pressure (mmHg)"], errors="coerce")
        if systolic_only and "Peripheral Diastolic Pressure (mmHg)" in df
        else None
    )

    min_distance = float("inf")
    min_diastolic_diff = float("inf")
    closest_pair: tuple[int, int] | None = None

    for i, idx_i in enumerate(df.index[:-1]):
        for idx_j in df.index[i + 1 :]:
            diff = df.loc[idx_i, fields] - df.loc[idx_j, fields]
            distance = (diff.pow(2).sum()) ** 0.5
            diastolic_diff = float("inf")
            if systolic_only and diastolic_values is not None:
                diastolic_diff = diastolic_values.loc[idx_i] - diastolic_values.loc[idx_j]
                diastolic_diff = (
                    abs(diastolic_diff) if pd.notna(diastolic_diff) else float("inf")
                )

            if distance < min_distance:
                min_distance = distance
                min_diastolic_diff = diastolic_diff
                closest_pair = (idx_i, idx_j)
            elif distance == min_distance and systolic_only:
                if diastolic_diff < min_diastolic_diff:
                    min_diastolic_diff = diastolic_diff
                    closest_pair = (idx_i, idx_j)

    return closest_pair


def average_pair_rows(
    pair_df: pd.DataFrame,
    excluded_fields: set[str],
) -> dict[str, object]:
    averaged: dict[str, object] = {}
    for column in pair_df.columns:
        if column in excluded_fields:
            continue
        if column == "Patient ID":
            averaged[column] = pair_df[column].iloc[0]
            continue

        numeric_values = pd.to_numeric(pair_df[column], errors="coerce")
        if numeric_values.notna().any():
            averaged[column] = numeric_values.mean()
        else:
            non_null = pair_df[column].dropna()
            averaged[column] = non_null.iloc[0] if not non_null.empty else None

    return averaged


def calculate_pair_differences(pair_df: pd.DataFrame) -> dict[str, float | None]:
    differences: dict[str, float | None] = {}
    for source_field, export_column in PAIR_DIFF_EXPORT_COLUMNS.items():
        if source_field not in pair_df.columns or len(pair_df.index) < 2:
            differences[export_column] = None
            continue

        numeric_values = pd.to_numeric(pair_df[source_field], errors="coerce")
        if len(numeric_values.index) < 2 or numeric_values.isna().any():
            differences[export_column] = None
            continue

        differences[export_column] = abs(
            float(numeric_values.iloc[0]) - float(numeric_values.iloc[1])
        )
    return differences


def pair_alert_triggered(
    pair_df: pd.DataFrame,
    threshold: float,
) -> bool:
    numeric_fields = [
        "Peripheral Systolic Pressure (mmHg)",
        "Peripheral Diastolic Pressure (mmHg)",
    ]
    for field in numeric_fields:
        if field not in pair_df.columns or len(pair_df.index) < 2:
            continue
        numeric_values = pd.to_numeric(pair_df[field], errors="coerce")
        if len(numeric_values.index) < 2 or numeric_values.isna().any():
            continue
        if abs(float(numeric_values.iloc[0]) - float(numeric_values.iloc[1])) > threshold:
            return True
    return False


def build_analyzed_data(
    df: pd.DataFrame,
    mode: int,
    manual_pairs: dict[str, tuple[int, int]] | None = None,
    pair_alert_threshold: float = 6.0,
) -> tuple[
    pd.DataFrame,
    set[int],
    dict[str, tuple[int, int]],
    dict[str, tuple[int, int]],
]:
    analysis_fields = ANALYSIS_FIELDS_BY_MODE.get(mode, ANALYSIS_FIELDS_BY_MODE[1])

    numeric_df = df.copy()
    for field in analysis_fields:
        numeric_df[field] = pd.to_numeric(numeric_df[field], errors="coerce")

    analyzed_records: list[dict[str, object]] = []
    kept_indices: set[int] = set()
    used_pairs: dict[str, tuple[int, int]] = {}
    auto_pairs: dict[str, tuple[int, int]] = {}
    excluded_fields = {
        "Source File",
        "Scanned ID",
        "Scan Date",
        "Scan Time",
        "Analyed",
        "Record #",
        "Source Path",
        "Subject ID",
        "Visit",
        "Timepoint",
        "Report Type",
    }

    manual_pairs = manual_pairs or {}

    for patient_id, group in numeric_df.groupby("Patient ID"):
        valid_group = group.dropna(subset=analysis_fields)
        auto_pair = closest_pair_indices(valid_group, analysis_fields)
        if auto_pair is not None:
            auto_pairs[patient_id] = auto_pair

        pair: tuple[int, int] | None = manual_pairs.get(patient_id)
        if not pair or not all(index in valid_group.index for index in pair):
            pair = auto_pair
        if pair is None:
            continue

        pair_df = df.loc[list(pair)]
        averaged_record = average_pair_rows(pair_df, excluded_fields)
        averaged_record.update(calculate_pair_differences(pair_df))
        averaged_record["Patient Entry Count"] = len(group)
        averaged_record["Pair Alert Threshold (mmHg)"] = pair_alert_threshold
        averaged_record["Pair Alert"] = (
            "Yes" if pair_alert_triggered(pair_df, pair_alert_threshold) else "No"
        )
        averaged_record["Patient ID"] = patient_id
        # Preserve the first row's Subject/Visit/Timepoint/Report Type for context.
        for context_field in ("Subject ID", "Visit", "Timepoint", "Report Type"):
            if context_field in pair_df.columns:
                value = pair_df[context_field].dropna()
                if not value.empty:
                    averaged_record[context_field] = value.iloc[0]
        analyzed_records.append(averaged_record)
        kept_indices.update(pair)
        used_pairs[patient_id] = pair

    return pd.DataFrame(analyzed_records), kept_indices, used_pairs, auto_pairs


def _compute_review_items(
    dataframe: pd.DataFrame,
    used_pairs: dict[str, tuple[int, int]],
    manual_patients: list[str],
    pair_alert_threshold: float,
) -> list[ReviewItem]:
    """Combine multi-entry and pair-alert review reasons into one ordered list."""
    entry_counts = patient_entry_counts(dataframe)

    pair_alert_patients: list[str] = []
    for patient_id, pair in used_pairs.items():
        if entry_counts.get(patient_id) != 2:
            continue
        pair_df = dataframe.loc[list(pair)]
        if pair_alert_triggered(pair_df, pair_alert_threshold):
            pair_alert_patients.append(patient_id)

    items: list[ReviewItem] = []
    for patient_id in manual_patients:
        is_alert = pair_alert_triggered(
            dataframe.loc[list(used_pairs[patient_id])],
            pair_alert_threshold,
        ) if patient_id in used_pairs else False
        reason = REVIEW_REASON_BOTH if is_alert else REVIEW_REASON_MULTI_ENTRY
        items.append(ReviewItem(patient_id=patient_id, reason=reason))

    for patient_id in pair_alert_patients:
        if any(item.patient_id == patient_id for item in items):
            continue
        items.append(ReviewItem(patient_id=patient_id, reason=REVIEW_REASON_PAIR_ALERT))

    return items


def build_analysis(
    records: list[dict[str, object]],
    manual_pairs: dict[str, tuple[int, int]] | None = None,
    mode: int = ANALYSIS_MODE,
    pair_alert_threshold: float = 6.0,
) -> AnalysisBundle:
    dataframe, special_row_mask = prepare_dataframe(records)
    analyzed_df, kept_indices, used_pairs, auto_pairs = build_analyzed_data(
        dataframe,
        mode,
        manual_pairs,
        pair_alert_threshold=pair_alert_threshold,
    )

    manual_patients = [
        patient_id
        for patient_id, group in dataframe.loc[dataframe["Special Row"] != True].groupby("Patient ID")
        if len(group) > 2
    ]

    review_items = _compute_review_items(
        dataframe,
        used_pairs,
        manual_patients,
        pair_alert_threshold,
    )

    return AnalysisBundle(
        dataframe=dataframe,
        special_row_mask=special_row_mask,
        analyzed_df=analyzed_df,
        kept_indices=kept_indices,
        used_pairs=used_pairs,
        auto_pairs=auto_pairs,
        manual_patients=manual_patients,
        review_items=review_items,
    )


def patient_entry_counts(df: pd.DataFrame) -> dict[str, int]:
    regular_rows = df.loc[df["Special Row"] != True]
    return {
        patient_id: int(len(group))
        for patient_id, group in regular_rows.groupby("Patient ID")
    }


def display_dataframe(bundle: AnalysisBundle) -> pd.DataFrame:
    frame = bundle.dataframe.copy()
    frame["Analyed"] = "No"
    if bundle.kept_indices:
        frame.loc[frame.index.isin(bundle.kept_indices), "Analyed"] = "Yes"
    return frame


def patient_rows(df: pd.DataFrame, patient_id: str) -> pd.DataFrame:
    return df.loc[(df["Patient ID"] == patient_id) & (df["Special Row"] != True)]


def initial_manual_pairs(
    df: pd.DataFrame,
    auto_pairs: dict[str, tuple[int, int]],
    review_patients: list[str],
) -> dict[str, list[int]]:
    pairs: dict[str, list[int]] = {}
    for patient_id in review_patients:
        auto_pair = list(auto_pairs.get(patient_id, ()))
        patient_frame = patient_rows(df, patient_id)
        fallback = list(patient_frame.index[:2])
        pairs[patient_id] = auto_pair[:2] if len(auto_pair) == 2 else fallback
    return pairs


def data_sheet_path(data_sheet_folder: Path | None, patient_id: str) -> Path | None:
    if data_sheet_folder is None or not data_sheet_folder.exists():
        return None

    subject_prefix = re.split(r"[ _]", patient_id, maxsplit=1)[0].lower()
    for candidate in sorted(data_sheet_folder.glob("*.pdf")):
        if candidate.stem.lower().startswith(subject_prefix):
            return candidate
    return None


def format_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def format_pressure_triplet(sys: object, dia: object, mean: object) -> str:
    if pd.isna(sys) and pd.isna(dia) and pd.isna(mean):
        return "—"

    parts: list[str] = []
    if not pd.isna(sys) or not pd.isna(dia):
        left = format_value(sys) or "—"
        right = format_value(dia) or "—"
        parts.append(f"{left}/{right}")
    if not pd.isna(mean):
        parts.append(f"MAP {format_value(mean)}")
    if not parts:
        return "—"
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]} ({parts[1]})"


def record_status(patient_id: object) -> str:
    value = str(patient_id or "")
    if value == CLINICAL_REPORT_MESSAGE:
        return "Wrong type (Clinical)"
    if value == DETAILED_REPORT_MESSAGE:
        return "Wrong type (Detailed)"
    if value == UNRECOGNIZED_REPORT_MESSAGE:
        return "Unrecognized"
    return "Detailed report"


def columns_for_report_mode(report_mode: str) -> list[str]:
    if report_mode == REPORT_MODE_CLINICAL:
        return CLINICAL_WORKBOOK_COLUMNS
    return WORKBOOK_COLUMNS


def filter_columns_for_report_mode(frame: pd.DataFrame, report_mode: str) -> pd.DataFrame:
    base_columns = columns_for_report_mode(report_mode)
    ordered_columns = [column for column in base_columns if column in frame.columns]
    remaining_columns = [
        column
        for column in frame.columns
        if column not in ordered_columns and column not in WORKBOOK_COLUMNS
    ]
    return frame[[*ordered_columns, *remaining_columns]]


def skipped_files_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Source File",
        "Patient ID",
        "Subject",
        "Visit",
        "Timepoint",
        "Report type",
        "Scan date",
        "Reason",
    ]
    if frame.empty or "Patient ID" not in frame.columns:
        return pd.DataFrame(columns=columns)

    special_mask = (
        frame["Special Row"]
        if "Special Row" in frame.columns
        else pd.Series(False, index=frame.index)
    )
    regular_frame = frame.loc[special_mask != True].copy()
    skipped_frames: list[pd.DataFrame] = []

    if not regular_frame.empty:
        entry_counts = regular_frame.groupby("Patient ID")["Patient ID"].transform("size")
        single_file_rows = regular_frame.loc[entry_counts == 1].copy()
        if not single_file_rows.empty:
            single_file_rows["Reason"] = "Only one file uploaded"
            skipped_frames.append(single_file_rows)

    special_rows = frame.loc[special_mask == True].copy()
    if not special_rows.empty:
        special_rows["Reason"] = special_rows["Patient ID"].map(record_status)
        special_rows["Patient ID"] = ""
        skipped_frames.append(special_rows)

    if not skipped_frames:
        return pd.DataFrame(columns=columns)

    skipped = pd.concat(skipped_frames, ignore_index=True).rename(
        columns={
            "Subject ID": "Subject",
            "Report Type": "Report type",
            "Scan Date": "Scan date",
        }
    )
    return skipped.reindex(columns=columns)


def save_to_excel(
    records: list[dict[str, object]],
    output_path: Path,
    manual_pairs: dict[str, tuple[int, int]] | None = None,
    mode: int = ANALYSIS_MODE,
    pair_alert_threshold: float = 6.0,
    report_mode: str = REPORT_MODE_DETAILED,
) -> int:
    bundle = build_analysis(
        records,
        manual_pairs=manual_pairs,
        mode=mode,
        pair_alert_threshold=pair_alert_threshold,
    )
    df = display_dataframe(bundle)

    kept_df = df[df["Analyed"] == "Yes"].copy()
    averaged_df = bundle.analyzed_df.drop(columns=["Record #"], errors="ignore").copy()

    date_columns = ["Scan Date", "Date of Birth", "Scan date"]

    def normalize_dates(frame: pd.DataFrame) -> pd.DataFrame:
        for date_column in date_columns:
            if date_column not in frame.columns:
                continue

            parsed_dates = pd.to_datetime(
                frame[date_column],
                errors="coerce",
                dayfirst=True,
            )
            frame.loc[:, date_column] = parsed_dates
        return frame

    df = normalize_dates(df)
    kept_df = normalize_dates(kept_df)
    averaged_df = normalize_dates(averaged_df)
    skipped_df = skipped_files_dataframe(df)

    def strip_aux_columns(frame: pd.DataFrame) -> pd.DataFrame:
        return frame.drop(
            columns=["Special Row", *EXTRA_COLUMNS, *UI_CONTEXT_COLUMNS],
            errors="ignore",
        )

    def order_workbook_columns(frame: pd.DataFrame) -> pd.DataFrame:
        return filter_columns_for_report_mode(frame, report_mode)

    df_to_save = order_workbook_columns(strip_aux_columns(df.copy()))
    kept_df_to_save = order_workbook_columns(strip_aux_columns(kept_df.copy()))
    averaged_df_to_save = order_workbook_columns(strip_aux_columns(averaged_df.copy()))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_to_save.to_excel(writer, sheet_name="All Data", index=False)
        kept_df_to_save.to_excel(writer, sheet_name="Kept Data", index=False)
        averaged_df_to_save.to_excel(writer, sheet_name="Averaged Data", index=False)
        skipped_df.to_excel(writer, sheet_name="Skipped Files", index=False)

        header_alignment = Alignment(horizontal="left")
        center_alignment = Alignment(horizontal="center")
        left_alignment = Alignment(horizontal="left")
        sheet_frames = {
            "All Data": df_to_save,
            "Kept Data": kept_df_to_save,
            "Averaged Data": averaged_df_to_save,
            "Skipped Files": skipped_df,
        }

        for sheet_name, frame in sheet_frames.items():
            sheet = writer.book[sheet_name]
            left_aligned_col_indices = {
                frame.columns.get_loc(column) + 1
                for column in ("Source File", "Patient ID", "Reason")
                if column in frame.columns
            }

            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
            ):
                for cell in row:
                    if cell.row == 1:
                        cell.alignment = header_alignment
                    elif cell.column in left_aligned_col_indices:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment

            for date_column in date_columns:
                if date_column not in frame.columns:
                    continue

                date_col_index = frame.columns.get_loc(date_column) + 1

                for column_cells in sheet.iter_cols(
                    min_col=date_col_index,
                    max_col=date_col_index,
                    min_row=2,
                    max_row=sheet.max_row,
                ):
                    for date_cell in column_cells:
                        date_cell.number_format = "MM/DD/YY"

    return len(df)
